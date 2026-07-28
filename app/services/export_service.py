from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from datetime import UTC, date, datetime
from io import BytesIO
import re
from xml.sax.saxutils import escape

from flask import current_app
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models import (
    Assignment,
    AssignmentSource,
    DiagnosticRun,
    Holiday,
    ScheduleMonth,
    ScheduleMonthStatus,
    ScheduleVersion,
)
from app.services.assignment_codes import ASSIGNMENT_CODE_CATALOG, ASSIGNMENT_CODE_DEFINITIONS
from app.services.diagnostic_service import latest_run
from app.services.monthly_grid_builder import MonthlyGridCell, build_monthly_grid
from app.services.schedule_version_policy import ScheduleVersionPolicy
from app.services.service_code_catalog import COVERAGE_TARGETS


EXCEL_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIMETYPE = "application/pdf"


@dataclass(frozen=True)
class ExcelExportResult:
    stream: BytesIO
    filename: str
    mimetype: str = EXCEL_MIMETYPE


@dataclass(frozen=True)
class PdfExportResult:
    stream: BytesIO
    filename: str
    mimetype: str = PDF_MIMETYPE


class ScheduleExcelExportError(Exception):
    pass


class SchedulePdfExportError(Exception):
    pass


class ExportStyleCatalog:
    HEADER = "1F4E79"
    HEADER_TEXT = "FFFFFF"
    SUBHEADER = "D9EAF7"
    WEEKEND = "F3F4F6"
    HOLIDAY = "FFE8A3"
    OUTSIDE = "E5E7EB"
    MANUAL = "FDE68A"
    LOCKED = "E0F2FE"
    OVERRIDE = "FCE7F3"
    BORDER = "CBD5E1"

    CODE_COLORS = {
        "AT1": "DBEAFE",
        "AT2": "BFDBFE",
        "AT3": "93C5FD",
        "PO1": "DCFCE7",
        "PO2": "BBF7D0",
        "PO3": "86EFAC",
        "PT": "CCFBF1",
        "P": "E5E7EB",
        "R": "E9D5FF",
        "CR": "DDD6FE",
        "DS": "FEF3C7",
        "DC": "FDE68A",
        "FF": "FBCFE8",
        "FC": "FED7AA",
        "FR": "FECACA",
        "LF": "FCA5A5",
        "LP": "FECACA",
        "BM": "D1D5DB",
        "LC": "FECACA",
        "LN": "FECACA",
        "DIL": "E5E7EB",
        "TRIB": "E5E7EB",
        "INQ": "E5E7EB",
        "DCP": "FED7AA",
        "D24": "FEF3C7",
        "FORMACAO": "E0E7FF",
        "TIRO": "E0E7FF",
        "OUTRA": "E5E7EB",
    }

    thin_border = Border(
        left=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
        top=Side(style="thin", color=BORDER),
        bottom=Side(style="thin", color=BORDER),
    )

    @classmethod
    def fill_for_code(cls, code: str | None) -> PatternFill:
        return PatternFill("solid", fgColor=cls.CODE_COLORS.get(code or "", "FFFFFF"))


class ScheduleExcelExportService:
    def export_version(self, schedule_month: ScheduleMonth, version: ScheduleVersion) -> ExcelExportResult:
        if version.schedule_month_id != schedule_month.id:
            raise ScheduleExcelExportError("A versao indicada nao pertence ao mes selecionado.")
        if not ScheduleVersionPolicy(version).can_export():
            raise ScheduleExcelExportError("O estado da versao nao permite exportacao.")

        grid = build_monthly_grid(schedule_month, version=version)
        diagnostic_run = self._diagnostic_run_for_export(version)
        holidays_by_date = self._holidays_by_date(schedule_month)
        exported_at = datetime.now(UTC)

        workbook = Workbook()
        monthly_sheet = workbook.active
        monthly_sheet.title = "Escala Mensal"
        self._write_monthly_sheet(monthly_sheet, grid, holidays_by_date, exported_at)
        self._write_legend_sheet(workbook.create_sheet("Legenda"))
        self._write_summary_sheet(workbook.create_sheet("Resumo"), grid, holidays_by_date, exported_at)
        self._write_diagnostic_sheet(workbook.create_sheet("Diagnostico"), diagnostic_run)
        self._write_manual_changes_sheet_if_needed(workbook, version)
        self._set_workbook_properties(workbook, schedule_month, version, exported_at)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return ExcelExportResult(stream=stream, filename=self._filename(schedule_month, version))

    def _write_monthly_sheet(self, sheet, grid, holidays_by_date: dict[date, list[Holiday]], exported_at: datetime) -> None:
        version = grid.version
        state_label = _state_export_label(version)
        unit_name = current_app.config.get("UNIT_NAME", "Unidade nao definida")
        official_label = _official_label(grid.schedule_month, version)

        total_columns = 5 + len(grid.days)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        sheet.cell(row=1, column=1, value=_safe_text(unit_name))
        sheet.cell(row=1, column=1).font = Font(size=14, bold=True, color=ExportStyleCatalog.HEADER_TEXT)
        sheet.cell(row=1, column=1).fill = PatternFill("solid", fgColor=ExportStyleCatalog.HEADER)
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        title = f"Escala de Servico - {grid.schedule_month.month:02d}/{grid.schedule_month.year}"
        if version.is_operational_test:
            title = f"{title} - TESTE OPERACIONAL - NAO PUBLICAR"
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
        sheet.cell(row=2, column=1, value=_safe_text(title))
        sheet.cell(row=2, column=1).font = Font(size=13, bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        metadata = [
            ("Versao", f"V{version.version_number}"),
            ("Estado", state_label),
            ("Oficial", official_label),
            ("Exportado em", exported_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Revisao de conteudo", str(version.content_revision)),
            ("Revisao validada", _format_optional(version.validated_revision)),
            ("Validada em", _format_datetime(version.validated_at)),
            ("Publicada em", _format_datetime(version.published_at)),
            ("Encerrada em", _format_datetime(version.closed_at)),
        ]
        if version.is_operational_test:
            metadata.insert(3, ("Teste operacional", _operational_test_label(version)))
        row_number = 4
        for label, value in metadata:
            sheet.cell(row=row_number, column=1, value=label)
            sheet.cell(row=row_number, column=2, value=_safe_text(value))
            row_number += 1

        header_row = row_number + 1
        headers = ["Seccao", "Equipa", "Nome", "NIM", "Tipo"]
        for index, header in enumerate(headers, start=1):
            self._header_cell(sheet.cell(row=header_row, column=index, value=header))
        for offset, day in enumerate(grid.days, start=6):
            holiday_suffix = " F" if day.date in holidays_by_date else ""
            value = f"{day.day_number}\n{day.weekday_label}{holiday_suffix}"
            cell = sheet.cell(row=header_row, column=offset, value=value)
            self._header_cell(cell)
            if day.is_weekend:
                cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.WEEKEND)
            if day.date in holidays_by_date:
                cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.HOLIDAY)

        current_row = header_row + 1
        for grid_row in grid.rows:
            team_code = _team_code_from_group(grid_row.group_label)
            values = [
                grid_row.group_label,
                team_code,
                grid_row.military.name,
                grid_row.military.nim,
                grid_row.military.functional_type,
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=current_row, column=column, value=_safe_text(value))
                self._body_cell(cell)
            for offset, grid_cell in enumerate(grid_row.cells, start=6):
                value = _display_value(grid_cell)
                cell = sheet.cell(row=current_row, column=offset, value=_safe_text(value))
                self._body_cell(cell)
                self._style_schedule_cell(cell, grid_cell, holidays_by_date)
                comment = _cell_comment(grid_cell)
                if comment:
                    cell.comment = Comment(_safe_text(comment), "Escala")
            current_row += 1

        self._size_monthly_sheet(sheet, header_row, total_columns)

    def _write_legend_sheet(self, sheet) -> None:
        rows = [
            ("Codigo", "Designacao", "Categoria", "Descricao"),
            *(
                (
                    definition.code,
                    definition.label,
                    definition.category,
                    _legend_description(definition.code),
                )
                for definition in ASSIGNMENT_CODE_DEFINITIONS
            ),
        ]
        for row_number, row_values in enumerate(rows, start=1):
            for column, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=_safe_text(value))
                if row_number == 1:
                    self._header_cell(cell)
                else:
                    self._body_cell(cell)
                    if column == 1:
                        cell.fill = ExportStyleCatalog.fill_for_code(str(value))
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 22
        sheet.column_dimensions["D"].width = 70

    def _write_summary_sheet(self, sheet, grid, holidays_by_date: dict[date, list[Holiday]], exported_at: datetime) -> None:
        rows = [
            ("Resumo", ""),
            ("Mes", f"{grid.schedule_month.month:02d}/{grid.schedule_month.year}"),
            ("Versao", f"V{grid.version.version_number}"),
            ("Estado", grid.version.status),
            ("Teste operacional", _operational_test_label(grid.version)),
            ("Exportado em", exported_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Militares relevantes", len(grid.rows)),
            ("Dias no mes", len(grid.days)),
            ("Feriados configurados", len(holidays_by_date)),
            ("", ""),
            ("Totais por tipo", ""),
            *_counter_rows(_count_militaries_by_type(grid)),
            ("", ""),
            ("Totais por equipa/seccao", ""),
            *_counter_rows(_count_militaries_by_group(grid)),
            ("", ""),
            ("Totais por codigo visivel", ""),
            *_counter_rows(_count_visible_codes(grid)),
            ("", ""),
            ("Origem das atribuicoes", ""),
            *_counter_rows(_count_assignment_sources(grid)),
            ("", ""),
            ("Celulas bloqueadas", _count_cells(grid, lambda cell: cell.is_locked)),
            ("Celulas com override", _count_cells(grid, lambda cell: cell.has_override)),
            ("FF agendadas", _count_cells(grid, lambda cell: cell.primary_code == "FF")),
            ("FC agendadas", _count_cells(grid, lambda cell: cell.primary_code == "FC")),
            ("FR agendadas", _count_cells(grid, lambda cell: cell.primary_code == "FR")),
            ("Indisponibilidades visiveis", _count_cells(grid, lambda cell: cell.unavailability is not None)),
            ("", ""),
            ("Cobertura minima AT/PO", ""),
            *_coverage_rows(grid),
        ]
        for row_number, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_number, column=1, value=_safe_text(label))
            sheet.cell(row=row_number, column=2, value=_safe_text(value))
            for column in (1, 2):
                cell = sheet.cell(row=row_number, column=column)
                self._body_cell(cell)
                if value == "":
                    cell.font = Font(bold=True)
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 60

    def _write_diagnostic_sheet(self, sheet, diagnostic_run: DiagnosticRun | None) -> None:
        headers = ["Execucao", "Estado", "Nivel", "Categoria", "Codigo", "Data", "Militar", "Titulo", "Descricao"]
        for column, header in enumerate(headers, start=1):
            self._header_cell(sheet.cell(row=1, column=column, value=header))
        if diagnostic_run is None:
            sheet.cell(row=2, column=1, value="Sem diagnostico persistido para esta versao.")
            return
        for row_number, issue in enumerate(diagnostic_run.issues, start=2):
            values = [
                diagnostic_run.id,
                diagnostic_run.status,
                issue.level,
                issue.category,
                issue.code,
                issue.assignment_date.isoformat() if issue.assignment_date else "",
                issue.military.name if issue.military else "",
                issue.title,
                issue.description,
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=_safe_text(value))
                self._body_cell(cell)
        widths = [12, 14, 12, 18, 26, 14, 28, 40, 70]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    def _write_manual_changes_sheet_if_needed(self, workbook: Workbook, version: ScheduleVersion) -> None:
        assignments = (
            Assignment.query.filter_by(schedule_version_id=version.id, is_manual=True)
            .order_by(Assignment.assignment_date.asc(), Assignment.military_id.asc(), Assignment.id.asc())
            .all()
        )
        if not assignments:
            return
        sheet = workbook.create_sheet("Alteracoes Manuais")
        headers = ["Data", "Militar", "NIM", "Codigo", "Bloqueada", "Override", "Motivo", "Notas", "Criada em", "Atualizada em"]
        for column, header in enumerate(headers, start=1):
            self._header_cell(sheet.cell(row=1, column=column, value=header))
        for row_number, assignment in enumerate(assignments, start=2):
            values = [
                assignment.assignment_date.isoformat(),
                assignment.military.name,
                assignment.military.nim,
                assignment.code or "",
                "Sim" if assignment.is_locked else "Nao",
                "Sim" if assignment.has_override else "Nao",
                assignment.override_reason or "",
                assignment.notes or "",
                _format_datetime(assignment.created_at),
                _format_datetime(assignment.updated_at),
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_number, column=column, value=_safe_text(value))
                self._body_cell(cell)
        for index, width in enumerate([14, 30, 16, 12, 12, 12, 40, 45, 20, 20], start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    def _set_workbook_properties(self, workbook: Workbook, schedule_month: ScheduleMonth, version: ScheduleVersion, exported_at: datetime) -> None:
        workbook.properties.title = f"Escala {schedule_month.month:02d}/{schedule_month.year} V{version.version_number}"
        workbook.properties.subject = "Exportacao operacional da escala"
        workbook.properties.creator = "Escala de Servico"
        workbook.properties.created = exported_at.replace(tzinfo=None)

    def _diagnostic_run_for_export(self, version: ScheduleVersion) -> DiagnosticRun | None:
        if version.validated_diagnostic_run_id:
            return DiagnosticRun.query.get(version.validated_diagnostic_run_id)
        return latest_run(version.id)

    def _holidays_by_date(self, schedule_month: ScheduleMonth) -> dict[date, list[Holiday]]:
        start = date(schedule_month.year, schedule_month.month, 1)
        end = date(schedule_month.year, schedule_month.month, monthrange(schedule_month.year, schedule_month.month)[1])
        holidays = (
            Holiday.query.filter(
                Holiday.is_active.is_(True),
                Holiday.holiday_date >= start,
                Holiday.holiday_date <= end,
            )
            .order_by(Holiday.holiday_date.asc(), Holiday.scope.asc(), Holiday.id.asc())
            .all()
        )
        grouped: dict[date, list[Holiday]] = {}
        for holiday in holidays:
            grouped.setdefault(holiday.holiday_date, []).append(holiday)
        return grouped

    def _filename(self, schedule_month: ScheduleMonth, version: ScheduleVersion) -> str:
        state = {
            ScheduleMonthStatus.DRAFT.value: "Rascunho",
            ScheduleMonthStatus.VALIDATED.value: "Validada",
            ScheduleMonthStatus.PUBLISHED.value: "Oficial",
            ScheduleMonthStatus.CLOSED.value: "Encerrada",
        }.get(version.status, version.status)
        marker = "_Teste_Operacional" if version.is_operational_test else ""
        raw = f"Escala_{schedule_month.year}_{schedule_month.month:02d}_Versao_{version.version_number}_{state}{marker}.xlsx"
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", raw)

    def _header_cell(self, cell) -> None:
        cell.font = Font(bold=True, color=ExportStyleCatalog.HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ExportStyleCatalog.thin_border

    def _body_cell(self, cell) -> None:
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = ExportStyleCatalog.thin_border

    def _style_schedule_cell(self, cell, grid_cell: MonthlyGridCell, holidays_by_date: dict[date, list[Holiday]]) -> None:
        if grid_cell.is_outside_period:
            cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.OUTSIDE)
        elif grid_cell.date in holidays_by_date:
            cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.HOLIDAY)
        elif grid_cell.is_weekend:
            cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.WEEKEND)
        if grid_cell.primary_code:
            cell.fill = ExportStyleCatalog.fill_for_code(grid_cell.primary_code)
        if grid_cell.assignment and grid_cell.assignment.source == AssignmentSource.MANUAL.value:
            cell.font = Font(bold=True)
        if grid_cell.is_locked:
            cell.border = Border(
                left=Side(style="medium", color=ExportStyleCatalog.BORDER),
                right=Side(style="medium", color=ExportStyleCatalog.BORDER),
                top=Side(style="medium", color=ExportStyleCatalog.BORDER),
                bottom=Side(style="medium", color=ExportStyleCatalog.BORDER),
            )
        if grid_cell.has_override:
            cell.fill = PatternFill("solid", fgColor=ExportStyleCatalog.OVERRIDE)

    def _size_monthly_sheet(self, sheet, header_row: int, total_columns: int) -> None:
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=6)
        widths = [18, 12, 30, 14, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for index in range(6, total_columns + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 8
        sheet.row_dimensions[header_row].height = 34
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        sheet.print_title_rows = f"{header_row}:{header_row}"
        sheet.print_area = f"A1:{get_column_letter(total_columns)}{sheet.max_row}"
        sheet.oddHeader.center.text = "Escala de Servico"
        sheet.oddFooter.center.text = "Pagina &P de &N"


class SchedulePdfExportService:
    def export_version(self, schedule_month: ScheduleMonth, version: ScheduleVersion) -> PdfExportResult:
        if version.schedule_month_id != schedule_month.id:
            raise SchedulePdfExportError("A versao indicada nao pertence ao mes selecionado.")
        if not ScheduleVersionPolicy(version).can_export():
            raise SchedulePdfExportError("O estado da versao nao permite exportacao.")

        grid = build_monthly_grid(schedule_month, version=version)
        diagnostic_run = self._diagnostic_run_for_export(version)
        holidays_by_date = self._holidays_by_date(schedule_month)
        exported_at = datetime.now(UTC)
        stream = BytesIO()
        page_size = landscape(A3)
        margins = 0.45 * cm
        doc = SimpleDocTemplate(
            stream,
            pagesize=page_size,
            leftMargin=margins,
            rightMargin=margins,
            topMargin=1.25 * cm,
            bottomMargin=1.05 * cm,
            title=f"Escala {schedule_month.month:02d}/{schedule_month.year} V{version.version_number}",
            subject="Exportacao operacional da escala mensal em PDF A3",
            author="Escala de Servico",
            keywords=f"escala,{schedule_month.year},{schedule_month.month:02d},versao {version.version_number}",
            pageCompression=0,
        )
        styles = _pdf_styles()
        story = []
        story.extend(self._monthly_grid_story(grid, holidays_by_date, exported_at, styles, doc.width))
        story.append(PageBreak())
        story.extend(self._legend_story(styles, doc.width))
        story.append(PageBreak())
        story.extend(self._summary_story(grid, holidays_by_date, exported_at, styles, doc.width))
        story.append(PageBreak())
        story.extend(self._diagnostic_story(diagnostic_run, styles, doc.width))
        manual_story = self._manual_changes_story(version, styles, doc.width)
        if manual_story:
            story.append(PageBreak())
            story.extend(manual_story)

        page_meta = _PdfPageMeta(grid, exported_at)
        doc.build(
            story,
            canvasmaker=lambda *args, **kwargs: _NumberedCanvas(*args, page_meta=page_meta, **kwargs),
        )
        stream.seek(0)
        return PdfExportResult(stream=stream, filename=self._filename(schedule_month, version))

    def _monthly_grid_story(self, grid, holidays_by_date: dict[date, list[Holiday]], exported_at: datetime, styles, width: float) -> list:
        metadata = self._metadata_rows(grid, exported_at)
        story = [
            Paragraph(_pdf_text(current_app.config.get("UNIT_NAME", "Unidade nao definida")), styles["unit"]),
            Paragraph(_pdf_text(f"Escala de Servico - {grid.schedule_month.month:02d}/{grid.schedule_month.year}"), styles["title"]),
            _simple_table(metadata, [3.5 * cm, width - 3.5 * cm], styles),
            Spacer(1, 0.18 * cm),
        ]
        if grid.version.is_operational_test:
            story.insert(2, Paragraph("TESTE OPERACIONAL - NAO PUBLICAR", styles["warning"]))
        headers = ["Seccao", "Equipa", "Nome", "NIM", "Tipo"]
        day_headers = [
            f"{day.day_number}<br/>{_pdf_text(day.weekday_label)}{' F' if day.date in holidays_by_date else ''}"
            for day in grid.days
        ]
        table_data = [[_p(item, styles["grid_header"]) for item in [*headers, *day_headers]]]
        for row in grid.rows:
            base_values = [
                row.group_label,
                _team_code_from_group(row.group_label),
                row.military.name,
                row.military.nim,
                row.military.functional_type,
            ]
            data_row = [_p(value, styles["grid_body"]) for value in base_values]
            data_row.extend(_p(_pdf_cell_value(cell), styles["grid_cell"]) for cell in row.cells)
            table_data.append(data_row)

        fixed_widths = [1.9 * cm, 1.25 * cm, 4.1 * cm, 1.55 * cm, 1.75 * cm]
        day_width = max(0.62 * cm, (width - sum(fixed_widths)) / max(len(grid.days), 1))
        table = Table(table_data, colWidths=[*fixed_widths, *([day_width] * len(grid.days))], repeatRows=1, splitByRow=1)
        table.setStyle(self._monthly_table_style(grid, holidays_by_date))
        story.append(table)
        if grid.version.status == ScheduleMonthStatus.DRAFT.value:
            story.append(Spacer(1, 0.15 * cm))
            story.append(Paragraph("Esta versao e um rascunho e nao constitui a escala oficial.", styles["warning"]))
        return story

    def _legend_story(self, styles, width: float) -> list:
        rows = [["Codigo", "Designacao", "Categoria", "Descricao"]]
        rows.extend(
            [
                definition.code,
                definition.label,
                definition.category,
                _legend_description(definition.code),
            ]
            for definition in ASSIGNMENT_CODE_DEFINITIONS
        )
        rows.extend(
            [
                ["M", "Manual", "Indicador", "Celula introduzida ou preservada manualmente."],
                ["B", "Bloqueada", "Indicador", "Celula protegida contra regeneracao automatica."],
                ["O", "Override", "Indicador", "Regra ultrapassada com autorizacao e motivo."],
                ["*", "Ciclo subjacente", "Indicador", "Existe DS/DC calculado sob o codigo visivel."],
                ["F", "Feriado", "Calendario", "Dia assinalado como feriado configurado."],
            ]
        )
        story = [Paragraph("Legenda", styles["section"])]
        story.append(_styled_table(rows, [1.9 * cm, 4.2 * cm, 3.0 * cm, width - 9.1 * cm], styles))
        return story

    def _summary_story(self, grid, holidays_by_date: dict[date, list[Holiday]], exported_at: datetime, styles, width: float) -> list:
        summary_rows = [
            ("Mes", f"{grid.schedule_month.month:02d}/{grid.schedule_month.year}"),
            ("Versao", f"V{grid.version.version_number}"),
            ("Estado", _state_export_label(grid.version)),
            ("Oficial", _official_label(grid.schedule_month, grid.version)),
            ("Exportado em", exported_at.strftime("%Y-%m-%d %H:%M UTC")),
            ("Militares relevantes", len(grid.rows)),
            ("Dias no mes", len(grid.days)),
            ("Feriados configurados", len(holidays_by_date)),
            ("Celulas bloqueadas", _count_cells(grid, lambda cell: cell.is_locked)),
            ("Celulas com override", _count_cells(grid, lambda cell: cell.has_override)),
            ("Alteracoes manuais", _count_cells(grid, lambda cell: cell.assignment is not None and cell.assignment.is_manual)),
            ("Atribuicoes automaticas", _count_cells(grid, lambda cell: cell.assignment is not None and cell.assignment.source == AssignmentSource.SYSTEM.value)),
            ("Indisponibilidades visiveis", _count_cells(grid, lambda cell: cell.unavailability is not None)),
        ]
        if grid.version.is_operational_test:
            summary_rows.insert(4, ("Teste operacional", _operational_test_label(grid.version)))
        story = [Paragraph("Resumo", styles["section"])]
        story.append(_simple_table([(label, str(value)) for label, value in summary_rows], [5.2 * cm, width - 5.2 * cm], styles))
        story.append(Spacer(1, 0.2 * cm))
        counter_rows = [
            ("Totais por tipo", _count_militaries_by_type(grid)),
            ("Totais por equipa/seccao", _count_militaries_by_group(grid)),
            ("Totais por codigo visivel", _count_visible_codes(grid)),
            ("Origem das atribuicoes", _count_assignment_sources(grid)),
        ]
        for title, counts in counter_rows:
            story.append(Paragraph(_pdf_text(title), styles["subsection"]))
            story.append(_simple_table(_counter_rows(counts), [5.2 * cm, width - 5.2 * cm], styles))
            story.append(Spacer(1, 0.14 * cm))
        story.append(Paragraph("Cobertura diaria AT/PO", styles["subsection"]))
        coverage_rows = [["Data", "AT1", "AT2", "AT3", "PO1", "PO2", "PO3", "Obrigatorio", "Falta", "Cumpre"]]
        coverage_rows.extend(_coverage_detail_rows(grid))
        story.append(_styled_table(coverage_rows, [2.1 * cm, *([1.45 * cm] * 6), 2.4 * cm, 5.0 * cm, 2.0 * cm], styles))
        return story

    def _diagnostic_story(self, diagnostic_run: DiagnosticRun | None, styles, width: float) -> list:
        story = [Paragraph("Diagnostico", styles["section"])]
        if diagnostic_run is None:
            story.append(Paragraph("Sem diagnostico persistido para esta versao.", styles["body"]))
            return story
        summary = [
            ("Execucao", diagnostic_run.id),
            ("Estado", diagnostic_run.status),
            ("Erros", diagnostic_run.total_errors),
            ("Avisos", diagnostic_run.total_warnings),
            ("Informacoes", diagnostic_run.total_infos),
        ]
        story.append(_simple_table([(label, str(value)) for label, value in summary], [4.0 * cm, width - 4.0 * cm], styles))
        story.append(Spacer(1, 0.2 * cm))
        rows = [["Nivel", "Categoria", "Codigo", "Data", "Militar", "Titulo", "Descricao"]]
        for issue in diagnostic_run.issues:
            rows.append(
                [
                    issue.level,
                    issue.category,
                    issue.code,
                    issue.assignment_date.isoformat() if issue.assignment_date else "",
                    issue.military.name if issue.military else "",
                    issue.title,
                    issue.description,
                ]
            )
        story.append(_styled_table(rows, [1.8 * cm, 3.0 * cm, 3.4 * cm, 2.1 * cm, 4.0 * cm, 5.0 * cm, width - 19.3 * cm], styles))
        return story

    def _manual_changes_story(self, version: ScheduleVersion, styles, width: float) -> list:
        assignments = (
            Assignment.query.filter_by(schedule_version_id=version.id, is_manual=True)
            .order_by(Assignment.assignment_date.asc(), Assignment.military_id.asc(), Assignment.id.asc())
            .all()
        )
        if not assignments:
            return []
        rows = [["Data", "Militar", "NIM", "Codigo", "Bloqueada", "Override", "Motivo", "Notas"]]
        for assignment in assignments:
            rows.append(
                [
                    assignment.assignment_date.isoformat(),
                    assignment.military.name,
                    assignment.military.nim,
                    assignment.code or "",
                    "Sim" if assignment.is_locked else "Nao",
                    "Sim" if assignment.has_override else "Nao",
                    assignment.override_reason or "",
                    assignment.notes or "",
                ]
            )
        return [
            Paragraph("Alteracoes Manuais", styles["section"]),
            _styled_table(rows, [2.1 * cm, 4.2 * cm, 2.0 * cm, 1.6 * cm, 2.0 * cm, 2.0 * cm, 5.0 * cm, width - 18.9 * cm], styles),
        ]

    def _monthly_table_style(self, grid, holidays_by_date: dict[date, list[Holiday]]) -> TableStyle:
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), _rl_hex(ExportStyleCatalog.HEADER)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, _rl_hex(ExportStyleCatalog.BORDER)),
            ("FONTNAME", (0, 1), (4, -1), "Helvetica"),
            ("ALIGN", (2, 1), (2, -1), "LEFT"),
            ("FONTSIZE", (0, 1), (-1, -1), 5.2),
            ("TOPPADDING", (0, 0), (-1, -1), 1.4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.4),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.2),
        ]
        for day_index, day in enumerate(grid.days, start=5):
            if day.date in holidays_by_date:
                style_commands.append(("BACKGROUND", (day_index, 0), (day_index, 0), _rl_hex(ExportStyleCatalog.HOLIDAY)))
                continue
            if day.is_weekend:
                style_commands.append(("BACKGROUND", (day_index, 0), (day_index, 0), _rl_hex(ExportStyleCatalog.WEEKEND)))
        for row_index, row in enumerate(grid.rows, start=1):
            for column_index, cell in enumerate(row.cells, start=5):
                color_hex = _pdf_cell_color(cell, holidays_by_date)
                style_commands.append(("BACKGROUND", (column_index, row_index), (column_index, row_index), _rl_hex(color_hex)))
                if cell.assignment and cell.assignment.source == AssignmentSource.MANUAL.value:
                    style_commands.append(("FONTNAME", (column_index, row_index), (column_index, row_index), "Helvetica-Bold"))
                if cell.is_locked:
                    style_commands.append(("BOX", (column_index, row_index), (column_index, row_index), 1.0, _rl_hex("0F172A")))
        return TableStyle(style_commands)

    def _metadata_rows(self, grid, exported_at: datetime) -> list[tuple[str, str]]:
        version = grid.version
        rows = [
            ("Versao", f"V{version.version_number}"),
            ("Estado", _state_export_label(version)),
            ("Oficial", _official_label(grid.schedule_month, version)),
            ("Revisao de conteudo", str(version.content_revision)),
            ("Revisao validada", _format_optional(version.validated_revision)),
            ("Validada em", _format_datetime(version.validated_at)),
            ("Publicada em", _format_datetime(version.published_at)),
            ("Encerrada em", _format_datetime(version.closed_at)),
            ("Exportado em", exported_at.strftime("%Y-%m-%d %H:%M UTC")),
        ]
        if version.is_operational_test:
            rows.insert(3, ("Teste operacional", _operational_test_label(version)))
        return rows

    def _diagnostic_run_for_export(self, version: ScheduleVersion) -> DiagnosticRun | None:
        if version.validated_diagnostic_run_id:
            return DiagnosticRun.query.get(version.validated_diagnostic_run_id)
        return latest_run(version.id)

    def _holidays_by_date(self, schedule_month: ScheduleMonth) -> dict[date, list[Holiday]]:
        return ScheduleExcelExportService()._holidays_by_date(schedule_month)

    def _filename(self, schedule_month: ScheduleMonth, version: ScheduleVersion) -> str:
        state = {
            ScheduleMonthStatus.DRAFT.value: "Rascunho",
            ScheduleMonthStatus.VALIDATED.value: "Validada",
            ScheduleMonthStatus.PUBLISHED.value: "Oficial",
            ScheduleMonthStatus.CLOSED.value: "Encerrada",
        }.get(version.status, version.status)
        marker = "_Teste_Operacional" if version.is_operational_test else ""
        raw = f"Escala_{schedule_month.year}_{schedule_month.month:02d}_Versao_{version.version_number}_{state}{marker}.pdf"
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", raw)


@dataclass(frozen=True)
class _PdfPageMeta:
    grid: object
    exported_at: datetime


class _NumberedCanvas(Canvas):
    def __init__(self, *args, page_meta: _PdfPageMeta, **kwargs):
        super().__init__(*args, **kwargs)
        self.page_meta = page_meta
        self._saved_page_states = []
        self.setTitle(f"Escala {page_meta.grid.schedule_month.month:02d}/{page_meta.grid.schedule_month.year} V{page_meta.grid.version.version_number}")
        self.setSubject("Exportacao operacional da escala mensal em PDF A3")
        self.setAuthor("Escala de Servico")
        self.setKeywords("escala,servico,pdf,a3")

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page_header_footer(page_count)
            super().showPage()
        super().save()

    def _draw_page_header_footer(self, page_count: int) -> None:
        grid = self.page_meta.grid
        version = grid.version
        width, height = self._pagesize
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        title = f"Escala de Servico {grid.schedule_month.month:02d}/{grid.schedule_month.year} - V{version.version_number} - {_state_export_label(version)}"
        self.drawString(0.45 * cm, height - 0.72 * cm, title[:150])
        self.setFont("Helvetica", 7)
        self.drawRightString(width - 0.45 * cm, height - 0.72 * cm, self.page_meta.exported_at.strftime("Exportado em %Y-%m-%d %H:%M UTC"))
        self.drawString(0.45 * cm, 0.45 * cm, "Documento operacional gerado pela aplicacao Escala de Servico")
        self.drawRightString(width - 0.45 * cm, 0.45 * cm, f"Pagina {self._pageNumber} de {page_count}")
        self.restoreState()


def _safe_text(value) -> str | int:
    if value is None:
        return ""
    if isinstance(value, int):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
    return text


def _format_optional(value) -> str:
    return "-" if value is None else str(value)


def _format_datetime(value) -> str:
    if value is None:
        return "-"
    return value.strftime("%Y-%m-%d %H:%M")


def _state_export_label(version: ScheduleVersion) -> str:
    if version.is_operational_test:
        return "TESTE OPERACIONAL - NAO PUBLICAR"
    labels = {
        ScheduleMonthStatus.DRAFT.value: "RASCUNHO - NAO OFICIAL",
        ScheduleMonthStatus.VALIDATED.value: "VALIDADA - AGUARDA PUBLICACAO",
        ScheduleMonthStatus.PUBLISHED.value: "VERSAO OFICIAL",
        ScheduleMonthStatus.CLOSED.value: "ESCALA ENCERRADA",
    }
    return labels.get(version.status, version.status)


def _official_label(schedule_month: ScheduleMonth, version: ScheduleVersion) -> str:
    if version.is_operational_test:
        return "Nao oficial - teste operacional"
    if schedule_month.published_version_id == version.id:
        return "Versao oficial publicada"
    return "Versao historica ou de trabalho"


def _operational_test_label(version: ScheduleVersion) -> str:
    if not version.is_operational_test:
        return "Nao"
    if version.is_archived:
        return "Sim - arquivado"
    return "Sim - nao publicar"


def _team_code_from_group(group_label: str) -> str:
    return group_label.replace("Equipa ", "") if group_label.startswith("Equipa ") else ""


def _display_value(cell: MonthlyGridCell) -> str:
    if cell.primary_code:
        return cell.primary_code
    if cell.team_code and not cell.is_outside_period:
        return cell.team_code
    return ""


def _cell_comment(cell: MonthlyGridCell) -> str:
    parts = []
    if cell.assignment and cell.assignment.is_visible:
        parts.append(f"Origem: {cell.assignment.source}")
        if cell.assignment.is_manual:
            parts.append("Alteracao manual")
        if cell.assignment.start_time and cell.assignment.end_time:
            parts.append(f"Horario: {cell.assignment.start_time.strftime('%H:%M')}-{cell.assignment.end_time.strftime('%H:%M')}")
        if cell.assignment.duration_minutes:
            parts.append(f"Duracao: {cell.assignment.duration_minutes} minutos")
        if cell.assignment.holiday_leave_credit:
            credit = cell.assignment.holiday_leave_credit
            parts.append(f"Credito FF #{credit.id} estado {credit.status}")
        if cell.assignment.compensatory_leave_credit:
            credit = cell.assignment.compensatory_leave_credit
            parts.append(f"Credito FC #{credit.id} estado {credit.status}")
        if cell.assignment.rescheduled_rest_credit:
            credit = cell.assignment.rescheduled_rest_credit
            parts.append(f"Credito FR #{credit.id} estado {credit.status}")
    if cell.cycle_code in {"DS", "DC"} and cell.primary_code != cell.cycle_code:
        parts.append(f"Ciclo subjacente: {cell.cycle_code}")
    if cell.unavailability:
        parts.append(f"Indisponibilidade: {cell.unavailability.code} {cell.unavailability.status}")
        if cell.is_partial_unavailability:
            parts.append("Indisponibilidade parcial")
    if cell.has_override:
        parts.append("Override autorizado")
        if cell.assignment and cell.assignment.override_reason:
            parts.append(f"Motivo: {cell.assignment.override_reason}")
    if cell.is_locked:
        parts.append("Celula bloqueada")
    if cell.notes:
        parts.append(f"Notas: {cell.notes}")
    if cell.warnings:
        parts.extend(cell.warnings)
    return " | ".join(parts)


def _legend_description(code: str) -> str:
    descriptions = {
        "DS": "Descanso semanal calculado pelo ciclo.",
        "DC": "Descanso complementar calculado pelo ciclo.",
        "FF": "Folga de feriado agendada e ligada a credito.",
        "FC": "Folga de compensacao agendada e ligada a credito.",
        "FR": "Folga reagendada agendada e ligada a credito.",
        "PT": "Patrulhamento adicional; nao conta para cobertura minima AT/PO.",
        "R": "Ronda manual; geracao automatica fora desta versao.",
        "CR": "Compensacao de ronda manual; geracao automatica fora desta versao.",
    }
    return descriptions.get(code, ASSIGNMENT_CODE_CATALOG.get(code).label if code in ASSIGNMENT_CODE_CATALOG else "")


def _count_militaries_by_type(grid) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in grid.rows:
        counts[row.military.functional_type] = counts.get(row.military.functional_type, 0) + 1
    return counts


def _count_militaries_by_group(grid) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in grid.rows:
        counts[row.group_label] = counts.get(row.group_label, 0) + 1
    return counts


def _count_visible_codes(grid) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in grid.rows:
        for cell in row.cells:
            if cell.primary_code:
                counts[cell.primary_code] = counts.get(cell.primary_code, 0) + 1
    return counts


def _count_assignment_sources(grid) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in grid.rows:
        for cell in row.cells:
            if cell.assignment and cell.assignment.is_visible:
                counts[cell.assignment.source] = counts.get(cell.assignment.source, 0) + 1
    return counts


def _count_cells(grid, predicate) -> int:
    return sum(1 for row in grid.rows for cell in row.cells if predicate(cell))


def _counter_rows(counts: dict[str, int]) -> list[tuple[str, str]]:
    if not counts:
        return [("Sem dados", "0")]
    return [(key, str(counts[key])) for key in sorted(counts)]


def _coverage_rows(grid) -> list[tuple[str, str]]:
    rows = []
    for day in grid.days:
        day_counts = {code: 0 for code in COVERAGE_TARGETS}
        for row in grid.rows:
            cell = row.cells[day.day_number - 1]
            if cell.primary_code in day_counts:
                day_counts[cell.primary_code] += 1
        missing = [
            f"{code} {day_counts[code]}/{target}"
            for code, target in COVERAGE_TARGETS.items()
            if day_counts[code] < target
        ]
        status = "OK" if not missing else "Incompleta: " + ", ".join(missing)
        rows.append((day.date.isoformat(), status))
    return rows


def _coverage_detail_rows(grid) -> list[list[str]]:
    rows = []
    for day in grid.days:
        day_counts = {code: 0 for code in COVERAGE_TARGETS}
        for row in grid.rows:
            cell = row.cells[day.day_number - 1]
            if cell.primary_code in day_counts:
                day_counts[cell.primary_code] += 1
        missing = [
            f"{code} {day_counts[code]}/{target}"
            for code, target in COVERAGE_TARGETS.items()
            if day_counts[code] < target
        ]
        rows.append(
            [
                day.date.isoformat(),
                *(str(day_counts[code]) for code in COVERAGE_TARGETS),
                str(sum(COVERAGE_TARGETS.values())),
                ", ".join(missing) if missing else "-",
                "Sim" if not missing else "Nao",
            ]
        )
    return rows


def _pdf_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "unit": ParagraphStyle("PdfUnit", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=12, leading=14, alignment=TA_CENTER, textColor=_rl_hex(ExportStyleCatalog.HEADER)),
        "title": ParagraphStyle("PdfTitle", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=10, leading=12, alignment=TA_CENTER, spaceAfter=5),
        "section": ParagraphStyle("PdfSection", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=14, spaceAfter=8),
        "subsection": ParagraphStyle("PdfSubsection", parent=base["Heading3"], fontName="Helvetica-Bold", fontSize=9, leading=11, spaceBefore=4, spaceAfter=4),
        "body": ParagraphStyle("PdfBody", parent=base["Normal"], fontName="Helvetica", fontSize=8, leading=10),
        "warning": ParagraphStyle("PdfWarning", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.darkred),
        "grid_header": ParagraphStyle("PdfGridHeader", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=5.5, leading=6.2, alignment=TA_CENTER, textColor=colors.white),
        "grid_body": ParagraphStyle("PdfGridBody", parent=base["Normal"], fontName="Helvetica", fontSize=5.3, leading=6.0, alignment=TA_LEFT),
        "grid_cell": ParagraphStyle("PdfGridCell", parent=base["Normal"], fontName="Helvetica", fontSize=4.8, leading=5.2, alignment=TA_CENTER),
        "table_header": ParagraphStyle("PdfTableHeader", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=7, leading=8, alignment=TA_CENTER, textColor=colors.white),
        "table_body": ParagraphStyle("PdfTableBody", parent=base["Normal"], fontName="Helvetica", fontSize=7, leading=8, alignment=TA_LEFT),
    }


def _p(value, style) -> Paragraph:
    return Paragraph(_pdf_text(value), style)


def _pdf_text(value) -> str:
    safe = _safe_text(value)
    text = str(safe)
    return escape(text)[:1000].replace("\n", "<br/>")


def _simple_table(rows: list[tuple[str, str]], widths: list[float], styles) -> Table:
    data = [[_p(label, styles["table_body"]), _p(value, styles["table_body"])] for label, value in rows]
    table = Table(data, colWidths=widths, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.25, _rl_hex(ExportStyleCatalog.BORDER)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (0, -1), _rl_hex(ExportStyleCatalog.SUBHEADER)),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _styled_table(rows: list[list[str]], widths: list[float], styles) -> Table:
    data = []
    for row_index, row in enumerate(rows):
        style = styles["table_header"] if row_index == 0 else styles["table_body"]
        data.append([_p(value, style) for value in row])
    table = Table(data, colWidths=widths, repeatRows=1, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _rl_hex(ExportStyleCatalog.HEADER)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, _rl_hex(ExportStyleCatalog.BORDER)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _pdf_cell_value(cell: MonthlyGridCell) -> str:
    value = _display_value(cell)
    indicators = []
    if cell.assignment and cell.assignment.is_visible:
        indicators.append("M" if cell.assignment.is_manual else "S")
    if cell.is_locked:
        indicators.append("B")
    if cell.has_override:
        indicators.append("O")
    if cell.cycle_code in {"DS", "DC"} and cell.primary_code != cell.cycle_code:
        indicators.append("*")
    if indicators:
        return f"{value}\n{' '.join(indicators)}" if value else " ".join(indicators)
    return value


def _pdf_cell_color(cell: MonthlyGridCell, holidays_by_date: dict[date, list[Holiday]]) -> str:
    if cell.has_override:
        return ExportStyleCatalog.OVERRIDE
    if cell.primary_code:
        return ExportStyleCatalog.CODE_COLORS.get(cell.primary_code, "FFFFFF")
    if cell.is_outside_period:
        return ExportStyleCatalog.OUTSIDE
    if cell.date in holidays_by_date:
        return ExportStyleCatalog.HOLIDAY
    if cell.is_weekend:
        return ExportStyleCatalog.WEEKEND
    return "FFFFFF"


def _rl_hex(value: str):
    return colors.HexColor(f"#{value}")
