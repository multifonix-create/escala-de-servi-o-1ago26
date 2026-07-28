from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    Assignment,
    AssignmentSource,
    DiagnosticCategory,
    DiagnosticIssue,
    DiagnosticLevel,
    DiagnosticRun,
    DiagnosticRunStatus,
    FunctionalType,
    Holiday,
    Military,
    MilitaryTeamHistory,
    ScheduleMonth,
    ScheduleMonthStatus,
    ScheduleVersion,
    ScheduleVersionSource,
    Team,
    TeamCycleReference,
)
from app.services.export_service import ScheduleExcelExportService


def _export_context(status=ScheduleMonthStatus.DRAFT.value, official=False):
    team = Team.query.filter_by(code="C").one()
    schedule_month = ScheduleMonth(year=2026, month=1, status=status)
    version = ScheduleVersion(
        schedule_month=schedule_month,
        version_number=1,
        status=status,
        source=ScheduleVersionSource.INITIAL.value,
        content_revision=3,
    )
    patrol = Military(
        name="=Militar Exportacao",
        nim="550001",
        functional_type=FunctionalType.PATRULHEIRO.value,
        start_date=date(2026, 1, 1),
    )
    sec = Military(
        name="Militar SEC",
        nim="550002",
        functional_type=FunctionalType.SEC.value,
        start_date=date(2026, 1, 1),
    )
    db.session.add_all([schedule_month, version, patrol, sec])
    db.session.flush()
    db.session.add(
        MilitaryTeamHistory(
            military_id=patrol.id,
            team_id=team.id,
            start_date=date(2026, 1, 1),
        )
    )
    db.session.add(
        TeamCycleReference(
            team_id=team.id,
            reference_date=date(2026, 1, 5),
            reference_phase=6,
            valid_from=date(2026, 1, 1),
        )
    )
    db.session.add(
        Holiday(
            holiday_date=date(2026, 1, 6),
            name="Feriado Exportacao",
            scope="NATIONAL",
            is_active=True,
        )
    )
    db.session.add_all(
        [
            Assignment(
                schedule_version_id=version.id,
                military_id=patrol.id,
                assignment_date=date(2026, 1, 5),
                code="AT1",
                source=AssignmentSource.MANUAL.value,
                is_manual=True,
                is_locked=True,
                has_override=True,
                override_reason="+Motivo protegido",
                notes="@Nota protegida",
                is_cleared=False,
            ),
            Assignment(
                schedule_version_id=version.id,
                military_id=sec.id,
                assignment_date=date(2026, 1, 6),
                code="PO1",
                source=AssignmentSource.SYSTEM.value,
                is_manual=False,
                is_locked=False,
                has_override=False,
                is_cleared=False,
            ),
        ]
    )
    run = DiagnosticRun(
        schedule_version_id=version.id,
        status=DiagnosticRunStatus.COMPLETED.value,
        total_errors=0,
        total_warnings=1,
        total_infos=0,
    )
    db.session.add(run)
    db.session.flush()
    db.session.add(
        DiagnosticIssue(
            diagnostic_run_id=run.id,
            level=DiagnosticLevel.WARNING.value,
            category=DiagnosticCategory.COVERAGE.value,
            code="TEST-EXPORT",
            title="Aviso de exportacao",
            description="Diagnostico persistido usado no Excel.",
            assignment_date=date(2026, 1, 5),
            military_id=patrol.id,
            is_blocking=False,
        )
    )
    if official:
        schedule_month.published_version_id = version.id
        version.published_at = datetime(2026, 1, 20, 9, 0)
    db.session.commit()
    return schedule_month, version


def _workbook_from_result(result):
    result.stream.seek(0)
    return load_workbook(result.stream)


def test_excel_export_creates_required_workbook_without_database_changes(app):
    with app.app_context():
        schedule_month, version = _export_context()
        counts_before = (
            Assignment.query.count(),
            DiagnosticRun.query.count(),
            DiagnosticIssue.query.count(),
        )

        result = ScheduleExcelExportService().export_version(schedule_month, version)
        workbook = _workbook_from_result(result)

        assert result.filename == "Escala_2026_01_Versao_1_Rascunho.xlsx"
        assert {"Escala Mensal", "Legenda", "Resumo", "Diagnostico", "Alteracoes Manuais"}.issubset(workbook.sheetnames)
        assert workbook["Escala Mensal"]["B5"].value == "RASCUNHO - NAO OFICIAL"
        assert workbook["Escala Mensal"]["C16"].value == "'=Militar Exportacao"
        assert workbook["Escala Mensal"]["J16"].value == "AT1"
        assert "Ciclo subjacente: DS" in workbook["Escala Mensal"]["J16"].comment.text
        assert workbook["Legenda"]["A2"].value == "AT1"
        assert "Cobertura minima AT/PO" in [cell.value for cell in workbook["Resumo"]["A"]]
        assert workbook["Diagnostico"]["E2"].value == "TEST-EXPORT"
        assert counts_before == (
            Assignment.query.count(),
            DiagnosticRun.query.count(),
            DiagnosticIssue.query.count(),
        )


def test_excel_export_route_downloads_xlsx(client, app):
    with app.app_context():
        schedule_month, version = _export_context(status=ScheduleMonthStatus.VALIDATED.value)
        url = f"/escala/{schedule_month.year}/{schedule_month.month}/versoes/{version.id}/exportar/excel"

    response = client.get(url)

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Escala_2026_01_Versao_1_Validada.xlsx" in response.headers["Content-Disposition"]
    workbook = load_workbook(BytesIO(response.data))
    assert workbook["Escala Mensal"]["B5"].value == "VALIDADA - AGUARDA PUBLICACAO"


def test_excel_export_identifies_official_published_version(app):
    with app.app_context():
        schedule_month, version = _export_context(status=ScheduleMonthStatus.PUBLISHED.value, official=True)

        workbook = _workbook_from_result(ScheduleExcelExportService().export_version(schedule_month, version))

        assert workbook["Escala Mensal"]["B5"].value == "VERSAO OFICIAL"
        assert workbook["Escala Mensal"]["B6"].value == "Versao oficial publicada"
